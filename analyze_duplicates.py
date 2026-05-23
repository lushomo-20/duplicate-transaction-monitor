from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\duplicate checker")
TRANSACTION_FILE = BASE_DIR / "TransactionList.xlsb"
SMARTCIT_FILE = BASE_DIR / "SmartCITReport.xlsm"
OUTPUT_FILE = BASE_DIR / "duplicate_report_june_2025_to_date.xlsx"
START_DATE = pd.Timestamp("2025-06-01")


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def clean_account(value: object) -> str:
    text = clean_text(value)
    text = text.lstrip("*").strip()
    return re.sub(r"\s+", " ", text)


def clean_vods(value: object) -> str:
    text = clean_text(value)
    match = re.search(r"\bVODS[A-Z0-9]+\b", text)
    return match.group(0) if match else text


def excel_date(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce").dt.normalize()
    numeric = pd.to_numeric(series, errors="coerce")
    converted = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    parsed = pd.to_datetime(series, errors="coerce")
    return converted.fillna(parsed).dt.normalize()


def excel_time(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce").dt.strftime("%H:%M:%S")
    numeric = pd.to_numeric(series, errors="coerce")
    seconds = (numeric * 86400).round().astype("Int64")
    time_from_number = pd.Series(pd.NA, index=series.index, dtype="object")
    valid = seconds.notna()
    time_from_number.loc[valid] = seconds.loc[valid].map(
        lambda s: f"{int(s // 3600) % 24:02d}:{int((s % 3600) // 60):02d}:{int(s % 60):02d}"
    )
    parsed = pd.to_datetime(series.astype(str), errors="coerce").dt.strftime("%H:%M:%S")
    return time_from_number.fillna(parsed).fillna("")


def pick_vods_column(df: pd.DataFrame) -> str | None:
    for col in ("TRACE_ID", "VODS_ID", "F"):
        if col in df.columns:
            values = df[col].dropna().astype(str).str.upper()
            if values.str.contains(r"\bVODS[A-Z0-9]+\b", regex=True).any():
                return col
    for col in df.columns:
        values = df[col].dropna().astype(str).str.upper()
        if values.str.contains(r"\bVODS[A-Z0-9]+\b", regex=True).any():
            return col
    return None


def load_transactions() -> pd.DataFrame:
    xl = pd.ExcelFile(TRANSACTION_FILE)
    frames: list[pd.DataFrame] = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(TRANSACTION_FILE, sheet_name=sheet, engine="pyxlsb")
        vods_col = pick_vods_column(df)
        if not vods_col or "ACCT_NAME" not in df.columns or "TRNS_AMT" not in df.columns:
            continue

        out = pd.DataFrame(
            {
                "source_sheet": sheet,
                "row_number": df.index + 2,
                "account_name": df["ACCT_NAME"],
                "account_key": df["ACCT_NAME"].map(clean_account),
                "vods_id": df[vods_col].map(clean_vods),
                "transaction_date": excel_date(df["EFF_DATE"] if "EFF_DATE" in df.columns else df["POST_DATE"]),
                "transaction_time": excel_time(df["TXN_TIME"]) if "TXN_TIME" in df.columns else "",
                "amount": pd.to_numeric(df["TRNS_AMT"], errors="coerce").round(2),
                "description": df["TRNS_DESC"] if "TRNS_DESC" in df.columns else "",
                "description_key": df["TRNS_DESC"].map(clean_text) if "TRNS_DESC" in df.columns else "",
                "ref_no": df["REF_NO"] if "REF_NO" in df.columns else "",
                "device": df["DEVICE"] if "DEVICE" in df.columns else "",
                "vods_source_column": vods_col,
            }
        )
        out = out[out["transaction_date"].ge(START_DATE)]
        out = out[out["vods_id"].str.startswith("VODS", na=False)]
        frames.append(out)

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def load_smartcit() -> pd.DataFrame:
    xl = pd.ExcelFile(SMARTCIT_FILE)
    candidates = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(SMARTCIT_FILE, sheet_name=sheet, engine="openpyxl")
        required = {"TRAN_DTE", "TRAN_TIME", "CLIENT_NAME", "AMOUNT", "AGENT_NAME", "VODS_ID"}
        if required.issubset(df.columns):
            candidates.append((sheet, df))
    if not candidates:
        return pd.DataFrame()

    sheet, df = max(candidates, key=lambda item: len(item[1]))
    out = pd.DataFrame(
        {
            "smartcit_sheet": sheet,
            "smartcit_row_number": df.index + 2,
            "smartcit_date": pd.to_datetime(df["TRAN_DTE"], errors="coerce").dt.normalize(),
            "smartcit_time": excel_time(df["TRAN_TIME"]),
            "smartcit_client_name": df["CLIENT_NAME"],
            "smartcit_client_key": df["CLIENT_NAME"].map(clean_account),
            "smartcit_amount": pd.to_numeric(df["AMOUNT"], errors="coerce").round(2),
            "agent_name": df["AGENT_NAME"],
            "agent_mobile": df["AGENT_MOBILE"] if "AGENT_MOBILE" in df.columns else "",
            "status": df["STATUS"] if "STATUS" in df.columns else "",
            "vods_id": df["VODS_ID"].map(clean_vods),
        }
    )
    out = out[out["smartcit_date"].ge(START_DATE)]
    out = out[out["vods_id"].str.startswith("VODS", na=False)]
    return out


def duplicate_rows(df: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if df.empty:
        return df
    counts = df.groupby(keys, dropna=False).size().rename("duplicate_count").reset_index()
    dup_keys = counts[counts["duplicate_count"].gt(1)]
    return df.merge(dup_keys, on=keys, how="inner").sort_values(keys + ["source_sheet", "row_number"])


def autosize_and_highlight(writer: pd.ExcelWriter, highlighted_sheets: set[str]) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    duplicate_fill = PatternFill("solid", fgColor="FFF2CC")

    for sheet_name, worksheet in writer.sheets.items():
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        if sheet_name in highlighted_sheets:
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = duplicate_fill
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def main() -> None:
    transactions = load_transactions()
    smartcit = load_smartcit()

    actual_key = ["account_key", "vods_id", "transaction_date", "transaction_time", "amount", "description_key"]
    possible_key = ["account_key", "transaction_date", "transaction_time", "amount", "description_key"]
    actual_duplicates = duplicate_rows(transactions, actual_key)
    possible_duplicates = duplicate_rows(transactions, possible_key)
    possible_duplicates = possible_duplicates[
        ~possible_duplicates.set_index(actual_key).index.isin(actual_duplicates.set_index(actual_key).index)
    ]

    smartcit_vods_duplicates = pd.DataFrame()
    smartcit_matched_agents = pd.DataFrame()
    smartcit_possible_duplicates = pd.DataFrame()
    smartcit_agent_ranking = pd.DataFrame()
    if not smartcit.empty:
        smartcit_vods_duplicates = smartcit.merge(
            smartcit.groupby("vods_id").size().rename("smartcit_vods_count").reset_index().query("smartcit_vods_count > 1"),
            on="vods_id",
            how="inner",
        ).sort_values(["vods_id", "smartcit_date", "smartcit_time", "agent_name"])

        smartcit_possible_key = ["smartcit_client_key", "smartcit_date", "smartcit_amount"]
        smartcit_possible_duplicates = smartcit.merge(
            smartcit.groupby(smartcit_possible_key, dropna=False)
            .size()
            .rename("smartcit_possible_count")
            .reset_index()
            .query("smartcit_possible_count > 1"),
            on=smartcit_possible_key,
            how="inner",
        ).sort_values(smartcit_possible_key + ["smartcit_time", "agent_name"])

        if not smartcit_possible_duplicates.empty:
            smartcit_agent_ranking = (
                smartcit_possible_duplicates.groupby(["agent_name", "agent_mobile"], dropna=False)
                .agg(
                    possible_duplicate_rows=("vods_id", "size"),
                    unique_vods_ids=("vods_id", "nunique"),
                    clients=("smartcit_client_key", "nunique"),
                    total_amount=("smartcit_amount", "sum"),
                )
                .reset_index()
                .sort_values(["possible_duplicate_rows", "unique_vods_ids"], ascending=False)
            )

    if not actual_duplicates.empty and not smartcit.empty:
        smartcit_matched_agents = actual_duplicates.merge(
            smartcit,
            on="vods_id",
            how="left",
            suffixes=("", "_smartcit"),
        ).sort_values(["vods_id", "source_sheet", "row_number", "agent_name"])

    duplicate_track_ids_by_sheet = pd.DataFrame()
    if not actual_duplicates.empty:
        duplicate_track_ids_by_sheet = (
            actual_duplicates.groupby("vods_id", dropna=False)
            .agg(
                duplicate_rows=("vods_id", "size"),
                source_sheets=("source_sheet", lambda values: ", ".join(sorted(set(map(str, values))))),
                accounts=("account_key", lambda values: ", ".join(sorted(set(map(str, values))))),
                first_date=("transaction_date", "min"),
                last_date=("transaction_date", "max"),
                total_amount=("amount", "sum"),
                devices=("device", lambda values: ", ".join(sorted(set(map(str, values))))),
            )
            .reset_index()
            .sort_values(["duplicate_rows", "vods_id"], ascending=[False, True])
        )

    top_accounts = (
        actual_duplicates.groupby(["account_key", "account_name"], dropna=False)
        .agg(duplicate_rows=("vods_id", "size"), duplicate_trace_ids=("vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_trace_ids"], ascending=False)
    )
    top_devices = (
        actual_duplicates.groupby("device", dropna=False)
        .agg(duplicate_rows=("vods_id", "size"), duplicate_trace_ids=("vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_trace_ids"], ascending=False)
    )
    top_amounts = (
        actual_duplicates.groupby("amount", dropna=False)
        .agg(duplicate_rows=("vods_id", "size"), duplicate_trace_ids=("vods_id", "nunique"), accounts=("account_key", "nunique"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_trace_ids"], ascending=False)
    )
    duplicate_trend = (
        actual_duplicates.groupby("transaction_date", dropna=False)
        .agg(duplicate_rows=("vods_id", "size"), duplicate_trace_ids=("vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values("transaction_date")
    )
    agent_ranking = pd.DataFrame()
    if not smartcit_matched_agents.empty:
        matched_with_agents = smartcit_matched_agents[smartcit_matched_agents["agent_name"].notna()]
        agent_ranking = (
            matched_with_agents.groupby(["agent_name", "agent_mobile"], dropna=False)
            .agg(
                duplicate_rows=("vods_id", "size"),
                duplicate_trace_ids=("vods_id", "nunique"),
                accounts=("account_key", "nunique"),
                devices=("device", "nunique"),
                total_amount=("amount", "sum"),
            )
            .reset_index()
            .sort_values(["duplicate_rows", "duplicate_trace_ids"], ascending=False)
        )

    direct_vods_overlap = 0
    if not transactions.empty and not smartcit.empty:
        direct_vods_overlap = len(set(actual_duplicates["vods_id"].dropna()) & set(smartcit["vods_id"].dropna()))

    summary = pd.DataFrame(
        [
            {"metric": "run_date", "value": date.today().isoformat()},
            {"metric": "date_filter_start", "value": START_DATE.date().isoformat()},
            {"metric": "transaction_rows_after_filter", "value": len(transactions)},
            {"metric": "smartcit_rows_after_filter", "value": len(smartcit)},
            {"metric": "actual_duplicate_rows", "value": len(actual_duplicates)},
            {"metric": "actual_duplicate_groups", "value": actual_duplicates.groupby(actual_key).ngroups if not actual_duplicates.empty else 0},
            {"metric": "actual_duplicate_track_vods_ids", "value": actual_duplicates["vods_id"].nunique() if not actual_duplicates.empty else 0},
            {"metric": "possible_duplicate_rows", "value": len(possible_duplicates)},
            {"metric": "possible_duplicate_groups", "value": possible_duplicates.groupby(possible_key).ngroups if not possible_duplicates.empty else 0},
            {"metric": "duplicate_track_vods_ids_found_in_smartcit_upload", "value": direct_vods_overlap},
            {"metric": "smartcit_duplicate_vods_rows", "value": len(smartcit_vods_duplicates)},
            {"metric": "actual_duplicate_rows_with_smartcit_agents", "value": len(smartcit_matched_agents[smartcit_matched_agents["agent_name"].notna()]) if not smartcit_matched_agents.empty else 0},
            {"metric": "smartcit_possible_duplicate_rows_same_client_date_amount", "value": len(smartcit_possible_duplicates)},
            {"metric": "smartcit_possible_duplicate_groups_same_client_date_amount", "value": smartcit_possible_duplicates.groupby(["smartcit_client_key", "smartcit_date", "smartcit_amount"]).ngroups if not smartcit_possible_duplicates.empty else 0},
        ]
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        actual_duplicates.to_excel(writer, index=False, sheet_name="Actual duplicates")
        possible_duplicates.to_excel(writer, index=False, sheet_name="Possible duplicates")
        smartcit_vods_duplicates.to_excel(writer, index=False, sheet_name="SmartCIT VODS duplicates")
        smartcit_possible_duplicates.to_excel(writer, index=False, sheet_name="SmartCIT possible duplicates")
        smartcit_matched_agents.to_excel(writer, index=False, sheet_name="Matched SmartCIT agents")
        duplicate_track_ids_by_sheet.to_excel(writer, index=False, sheet_name="2607 duplicated IDs by sheet")
        top_accounts.to_excel(writer, index=False, sheet_name="Top duplicated accounts")
        top_devices.to_excel(writer, index=False, sheet_name="Top duplicate devices")
        top_amounts.to_excel(writer, index=False, sheet_name="Top duplicate amounts")
        duplicate_trend.to_excel(writer, index=False, sheet_name="Duplicate trend by day")
        agent_ranking.to_excel(writer, index=False, sheet_name="Matched agent ranking")
        smartcit_agent_ranking.to_excel(writer, index=False, sheet_name="SmartCIT agent ranking")
        transactions.to_excel(writer, index=False, sheet_name="Transactions filtered")
        smartcit.to_excel(writer, index=False, sheet_name="SmartCIT filtered")
        autosize_and_highlight(
            writer,
            {
                "Actual duplicates",
                "Possible duplicates",
                "SmartCIT VODS duplicates",
                "SmartCIT possible duplicates",
                "Matched SmartCIT agents",
                "2607 duplicated IDs by sheet",
            },
        )

    print(summary.to_string(index=False))
    print(f"Report written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
