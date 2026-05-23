from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"C:\duplicate checker")
TRANSACTION_FILE = BASE_DIR / "TransactionList.xlsb"
OUTPUT_FILE = BASE_DIR / "transaction_description_duplicate_report_june_2025_to_date.xlsx"
START_DATE = pd.Timestamp("2025-06-01")


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def clean_account(value: object) -> str:
    return clean_text(value).lstrip("*").strip()


def clean_vods(value: object) -> str:
    text = clean_text(value)
    match = re.search(r"\bVODS[A-Z0-9]+\b", text)
    return match.group(0) if match else text


def compact_text(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean_text(value))


def description_category(*values: object) -> str:
    joined = " ".join(clean_text(value) for value in values if not pd.isna(value))
    compact = compact_text(joined)
    if "FNBOBPMT" in compact:
        return "FNB OB PMT"
    if "SMARTCIT" in compact:
        return "Smart CIT"
    if "CASHDEPOSITPARTNER" in compact:
        return "Cash deposit partner"
    return ""


def excel_date(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    converted = pd.to_datetime(numeric, unit="D", origin="1899-12-30", errors="coerce")
    parsed = pd.to_datetime(series, errors="coerce")
    return converted.fillna(parsed).dt.normalize()


def excel_time(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    seconds = (numeric * 86400).round().astype("Int64")
    out = pd.Series(pd.NA, index=series.index, dtype="object")
    valid = seconds.notna()
    out.loc[valid] = seconds.loc[valid].map(
        lambda s: f"{int(s // 3600) % 24:02d}:{int((s % 3600) // 60):02d}:{int(s % 60):02d}"
    )
    parsed = pd.to_datetime(series.astype(str), errors="coerce").dt.strftime("%H:%M:%S")
    return out.fillna(parsed).fillna("")


def pick_vods_column(df: pd.DataFrame) -> str | None:
    for col in ("TRACE_ID", "VODS_ID", "TRACK_ID", "TRACK ID", "F"):
        if col in df.columns:
            values = df[col].dropna().astype(str).str.upper()
            if values.str.contains(r"\bVODS[A-Z0-9]+\b", regex=True).any():
                return col
    for col in df.columns:
        values = df[col].dropna().astype(str).str.upper()
        if values.str.contains(r"\bVODS[A-Z0-9]+\b", regex=True).any():
            return col
    return None


def load_transactions() -> pd.DataFrame:
    xl = pd.ExcelFile(TRANSACTION_FILE)
    frames: list[pd.DataFrame] = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(TRANSACTION_FILE, sheet_name=sheet, engine="pyxlsb")
        vods_col = pick_vods_column(df)
        if not vods_col or "ACCT_NAME" not in df.columns or "TRNS_AMT" not in df.columns:
            continue

        trns_desc = df["TRNS_DESC"] if "TRNS_DESC" in df.columns else pd.Series("", index=df.index)
        desc_cont = df["TRNS_DESC_CONT"] if "TRNS_DESC_CONT" in df.columns else pd.Series("", index=df.index)
        trn_cde_desc = df["TRN_CDE_DESC"] if "TRN_CDE_DESC" in df.columns else pd.Series("", index=df.index)
        sub_cat = df["SUB_CAT1_DESC"] if "SUB_CAT1_DESC" in df.columns else pd.Series("", index=df.index)

        out = pd.DataFrame(
            {
                "source_sheet": sheet,
                "row_number": df.index + 2,
                "account_name": df["ACCT_NAME"],
                "account_key": df["ACCT_NAME"].map(clean_account),
                "track_vods_id": df[vods_col].map(clean_vods),
                "transaction_date": excel_date(df["EFF_DATE"] if "EFF_DATE" in df.columns else df["POST_DATE"]),
                "transaction_time": excel_time(df["TXN_TIME"]) if "TXN_TIME" in df.columns else "",
                "amount": pd.to_numeric(df["TRNS_AMT"], errors="coerce").round(2),
                "description": trns_desc,
                "description_key": trns_desc.map(clean_text),
                "description_cont": desc_cont,
                "trn_cde_desc": trn_cde_desc,
                "sub_cat1_desc": sub_cat,
                "ref_no": df["REF_NO"] if "REF_NO" in df.columns else "",
                "device": df["DEVICE"] if "DEVICE" in df.columns else "",
                "track_vods_source_column": vods_col,
            }
        )
        out["description_category"] = [
            description_category(a, b, c, d)
            for a, b, c, d in zip(trns_desc, desc_cont, trn_cde_desc, sub_cat)
        ]
        out = out[out["transaction_date"].ge(START_DATE)]
        out = out[out["track_vods_id"].str.startswith("VODS", na=False)]
        out = out[out["description_category"].ne("")]
        frames.append(out)

    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def duplicate_rows(df: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if df.empty:
        return df
    counts = df.groupby(keys, dropna=False).size().rename("duplicate_count").reset_index()
    dup_keys = counts[counts["duplicate_count"].gt(1)]
    return df.merge(dup_keys, on=keys, how="inner").sort_values(keys + ["source_sheet", "row_number"])


def autosize_and_highlight(writer: pd.ExcelWriter, highlighted_sheets: set[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    duplicate_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet_name, worksheet in writer.sheets.items():
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        if sheet_name in highlighted_sheets:
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = duplicate_fill
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def main() -> None:
    transactions = load_transactions()
    actual_key = [
        "account_key",
        "track_vods_id",
        "transaction_date",
        "transaction_time",
        "amount",
        "description_category",
        "description_key",
    ]
    possible_key = ["account_key", "transaction_date", "transaction_time", "amount", "description_category", "description_key"]

    actual_duplicates = duplicate_rows(transactions, actual_key)
    possible_duplicates = duplicate_rows(transactions, possible_key)
    if not possible_duplicates.empty and not actual_duplicates.empty:
        possible_duplicates = possible_duplicates[
            ~possible_duplicates.set_index(actual_key).index.isin(actual_duplicates.set_index(actual_key).index)
        ]

    duplicated_ids_by_sheet = (
        actual_duplicates.groupby("track_vods_id", dropna=False)
        .agg(
            duplicate_rows=("track_vods_id", "size"),
            source_sheets=("source_sheet", lambda values: ", ".join(sorted(set(map(str, values))))),
            description_categories=("description_category", lambda values: ", ".join(sorted(set(map(str, values))))),
            accounts=("account_key", lambda values: ", ".join(sorted(set(map(str, values))))),
            first_date=("transaction_date", "min"),
            last_date=("transaction_date", "max"),
            total_amount=("amount", "sum"),
            devices=("device", lambda values: ", ".join(sorted(set(map(str, values))))),
        )
        .reset_index()
        .sort_values(["duplicate_rows", "track_vods_id"], ascending=[False, True])
        if not actual_duplicates.empty
        else pd.DataFrame()
    )

    top_accounts = (
        actual_duplicates.groupby(["account_key", "account_name"], dropna=False)
        .agg(duplicate_rows=("track_vods_id", "size"), duplicate_ids=("track_vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_ids"], ascending=False)
    )
    top_devices = (
        actual_duplicates.groupby("device", dropna=False)
        .agg(duplicate_rows=("track_vods_id", "size"), duplicate_ids=("track_vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_ids"], ascending=False)
    )
    top_amounts = (
        actual_duplicates.groupby("amount", dropna=False)
        .agg(duplicate_rows=("track_vods_id", "size"), duplicate_ids=("track_vods_id", "nunique"), accounts=("account_key", "nunique"))
        .reset_index()
        .sort_values(["duplicate_rows", "duplicate_ids"], ascending=False)
    )
    category_summary = (
        actual_duplicates.groupby("description_category", dropna=False)
        .agg(duplicate_rows=("track_vods_id", "size"), duplicate_ids=("track_vods_id", "nunique"), accounts=("account_key", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values("duplicate_rows", ascending=False)
    )
    duplicate_trend = (
        actual_duplicates.groupby(["transaction_date", "description_category"], dropna=False)
        .agg(duplicate_rows=("track_vods_id", "size"), duplicate_ids=("track_vods_id", "nunique"), total_amount=("amount", "sum"))
        .reset_index()
        .sort_values("transaction_date")
    )

    summary = pd.DataFrame(
        [
            {"metric": "run_date", "value": date.today().isoformat()},
            {"metric": "date_filter_start", "value": START_DATE.date().isoformat()},
            {"metric": "matching_transaction_rows_after_filter", "value": len(transactions)},
            {"metric": "actual_duplicate_rows", "value": len(actual_duplicates)},
            {"metric": "actual_duplicate_groups", "value": actual_duplicates.groupby(actual_key).ngroups if not actual_duplicates.empty else 0},
            {"metric": "actual_duplicate_track_vods_ids", "value": actual_duplicates["track_vods_id"].nunique() if not actual_duplicates.empty else 0},
            {"metric": "possible_duplicate_rows", "value": len(possible_duplicates)},
            {"metric": "possible_duplicate_groups", "value": possible_duplicates.groupby(possible_key).ngroups if not possible_duplicates.empty else 0},
        ]
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        category_summary.to_excel(writer, index=False, sheet_name="Summary by description")
        actual_duplicates.to_excel(writer, index=False, sheet_name="Actual duplicates")
        possible_duplicates.to_excel(writer, index=False, sheet_name="Possible duplicates")
        duplicated_ids_by_sheet.to_excel(writer, index=False, sheet_name="Duplicated IDs by sheet")
        top_accounts.to_excel(writer, index=False, sheet_name="Top duplicated accounts")
        top_devices.to_excel(writer, index=False, sheet_name="Top duplicate devices")
        top_amounts.to_excel(writer, index=False, sheet_name="Top duplicate amounts")
        duplicate_trend.to_excel(writer, index=False, sheet_name="Duplicate trend by day")
        transactions.to_excel(writer, index=False, sheet_name="Filtered transactions")
        autosize_and_highlight(writer, {"Actual duplicates", "Possible duplicates", "Duplicated IDs by sheet"})

    print(summary.to_string(index=False))
    print(f"Report written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
